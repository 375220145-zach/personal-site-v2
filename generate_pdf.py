"""Generate competitive-analysis-portfolio.pdf from Excel files."""
import openpyxl
from fpdf import FPDF
import re

# Paths
SPECIFIC = '/Users/mac/Desktop/竞品分析/具体/竞品分析_绿联67W超能块自带线20000mAh充电宝_2026-05-29.xlsx'
CONCEPT = '/Users/mac/Desktop/竞品分析/概念/竞品分析_无拉链行李箱（新型闭合方式）_2026-06-01.xlsx'
OUTPUT = '/Users/mac/Desktop/cc项目库/personal-site V2/public/competitive-analysis-portfolio.pdf'

SPECIFIC_SHEETS = ['竞品分析', 'SWOT', '维度评分', '市场空白与建议']
CONCEPT_SHEETS = ['相似产品清单', '功能对比矩阵', '冲突评估与建议']

FONT_PATH = '/System/Library/Fonts/STHeiti Light.ttc'

# Unicode symbols that STHeiti doesn't support → Chinese text equivalents
SYMBOL_MAP = {
    '✓': '有',
    '✗': '无',
    '🔴': '[高]',
    '🟡': '[中]',
    '⚠': '[!]',
    '⎓': '',       # technical symbol, just remove
    '★': '★',     # star — STHeiti has this, kept as-is in the map for clarity
    '✦': '>',
    '→': '→',     # arrow — STHeiti should have this
}

def sanitize_text(text: str) -> str:
    """Replace Unicode symbols unsupported by the Chinese font."""
    if not text:
        return ''
    result = str(text)
    for sym, replacement in SYMBOL_MAP.items():
        result = result.replace(sym, replacement)
    return result


class PDF(FPDF):
    def __init__(self):
        super().__init__('L', 'mm', 'A4')
        self.add_font('zh', '', FONT_PATH)
        self.add_font('zh', 'B', FONT_PATH)
        self.set_auto_page_break(True, 10)

    def header(self):
        if self.page_no() == 1:
            return
        self.set_font('zh', '', 7)
        self.set_text_color(120, 120, 120)
        self.cell(0, 4, 'Competitive Analysis Portfolio — Zachary Pan', align='R')
        self.ln(6)

    def footer(self):
        self.set_y(-8)
        self.set_font('zh', '', 6)
        self.set_text_color(150, 150, 150)
        self.cell(0, 6, str(self.page_no()), align='C')

    def section_title(self, title: str):
        self.set_font('zh', 'B', 13)
        self.set_text_color(30, 30, 30)
        self.cell(0, 8, title)
        self.ln(10)

    def sub_title(self, title: str):
        self.set_font('zh', 'B', 10)
        self.set_text_color(60, 60, 60)
        self.cell(0, 6, title)
        self.ln(7)

    def body_text(self, text: str):
        self.set_font('zh', '', 8)
        self.set_text_color(80, 80, 80)
        self.multi_cell(0, 4.5, sanitize_text(text))
        self.ln(1)

    def table(self, headers, rows, col_widths=None, font_size=6.5):
        """Draw a table with headers and rows. All text is sanitized for font compatibility."""
        ncols = len(headers)
        usable = self.w - 16

        if not col_widths:
            col_widths = [usable / ncols] * ncols

        # Sanitize headers and rows
        sane_headers = [sanitize_text(str(h)) for h in headers]
        sane_rows = []
        for row in rows:
            sane_rows.append([sanitize_text(str(c)) if c is not None else '' for c in row])

        # Header
        self.set_font('zh', 'B', min(7, font_size + 0.5))
        self.set_fill_color(245, 245, 245)
        self.set_text_color(40, 40, 40)
        x_start = self.get_x()
        for i, h in enumerate(sane_headers):
            self.set_xy(x_start + sum(col_widths[:i]), self.get_y())
            self.cell(col_widths[i], 6, h[:80], border=1, fill=True, align='C')
        self.ln()

        # Rows
        for row_idx, row in enumerate(sane_rows):
            # Calculate max height needed for this row
            max_lines = 1
            for i, text in enumerate(row):
                w = col_widths[i] - 1.2
                self.set_font('zh', '', font_size)
                text_w = self.get_string_width(text)
                lines = max(1, int(text_w / max(w, 1)) + 1)
                if len(text) > 100:
                    lines = max(lines, 4)
                if len(text) > 200:
                    lines = max(lines, 6)
                max_lines = max(max_lines, min(lines, 10))

            row_h = max(5, max_lines * (font_size * 0.55))

            # Page break if needed
            if self.get_y() + row_h > self.h - 15:
                self.add_page()
                # Reprint header on new page
                self.set_font('zh', 'B', min(7, font_size + 0.5))
                self.set_fill_color(245, 245, 245)
                self.set_text_color(40, 40, 40)
                for i, h in enumerate(sane_headers):
                    self.set_xy(x_start + sum(col_widths[:i]), self.get_y())
                    self.cell(col_widths[i], 6, h[:80], border=1, fill=True, align='C')
                self.ln()

            y_before = self.get_y()

            for i, text in enumerate(row):
                self.set_xy(x_start + sum(col_widths[:i]), y_before)
                bg = 252 if row_idx % 2 == 0 else 255
                self.set_fill_color(bg, bg, bg)
                self.set_font('zh', '', font_size)
                self.set_text_color(70, 70, 70)
                # Truncate very long text but keep enough for context
                display_text = text[:300]
                self.multi_cell(col_widths[i], font_size * 0.55, display_text, border='LR', fill=True, align='L')

            self.set_xy(x_start, y_before)
            self.ln(row_h)


def process_specific(pdf: PDF):
    wb = openpyxl.load_workbook(SPECIFIC)
    pdf.add_page()
    pdf.section_title('具体竞品分析 — 绿联 67W 充电宝竞品对比')
    pdf.body_text('针对已上市实体产品，搜索同品类 5 款竞品进行全维度横向对比。覆盖基础信息、评分口碑、销量排名、产品规格、SWOT、维度评分、市场空白与战略建议。')

    for sn in SPECIFIC_SHEETS:
        ws = wb[sn]
        pdf.add_page()
        pdf.sub_title(f'{sn}')

        all_rows = list(ws.iter_rows(values_only=True))
        rows = [r for r in all_rows if any(c is not None for c in r)]
        if not rows:
            continue

        headers = [str(c) if c else '' for c in rows[0]]
        data = rows[1:]
        if not data:
            continue

        ncols = len(headers)
        usable = pdf.w - 16
        if ncols <= 2:
            cw = [usable * 0.3, usable * 0.7]
        elif ncols <= 4:
            cw = [usable / ncols] * ncols
        else:
            first_w = usable * 0.14
            rest_w = (usable - first_w) / (ncols - 1)
            cw = [first_w] + [rest_w] * (ncols - 1)

        pdf.table(headers, data, cw)


def process_concept(pdf: PDF):
    wb = openpyxl.load_workbook(CONCEPT)
    pdf.add_page()
    pdf.section_title('概念竞品分析 — 无拉链行李箱闭合方案')
    pdf.body_text('针对创新概念（无拉链行李箱），搜索 8 个相似产品/众筹项目 + 2 个相关专利，构建功能对比矩阵，进行冲突等级判定和差异化路径建议。')

    for sn in CONCEPT_SHEETS:
        ws = wb[sn]
        pdf.add_page()
        pdf.sub_title(f'{sn}')

        all_rows = list(ws.iter_rows(values_only=True))
        rows = [r for r in all_rows if any(c is not None for c in r)]
        if not rows:
            continue

        headers = [str(c) if c else '' for c in rows[0]]
        data = rows[1:]
        if not data:
            continue

        ncols = len(headers)
        usable = pdf.w - 16

        # Concept sheets can be very wide (10+ columns) — use smaller font and tighter widths
        if ncols >= 8:
            # Very wide: use smaller font, first col wider for labels
            fs = 5.5
            first_w = usable * 0.10
            rest_w = (usable - first_w) / (ncols - 1)
            cw = [first_w] + [rest_w] * (ncols - 1)
        elif ncols <= 2:
            cw = [usable * 0.25, usable * 0.75]
            fs = 6.5
        elif ncols <= 4:
            cw = [usable / ncols] * ncols
            fs = 6.5
        else:
            first_w = usable * 0.12
            rest_w = (usable - first_w) / (ncols - 1)
            cw = [first_w] + [rest_w] * (ncols - 1)
            fs = 6

        pdf.table(headers, data, cw, font_size=fs)


def main():
    pdf = PDF()
    pdf.set_margin(8)

    # Cover page
    pdf.add_page()
    pdf.ln(40)
    pdf.set_font('zh', 'B', 22)
    pdf.set_text_color(30, 30, 30)
    pdf.cell(0, 12, '竞品分析案例报告', align='C')
    pdf.ln(14)
    pdf.set_font('zh', '', 11)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, 'Competitive Analysis Portfolio', align='C')
    pdf.ln(20)

    pdf.set_draw_color(200, 200, 200)
    pdf.line(60, pdf.get_y(), pdf.w - 60, pdf.get_y())
    pdf.ln(14)

    pdf.set_font('zh', '', 10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 7, '具体竞品分析：绿联 67W 超能块自带线 20000mAh 充电宝', align='C')
    pdf.ln(7)
    pdf.cell(0, 7, '概念竞品分析：无拉链行李箱（新型闭合方式）', align='C')
    pdf.ln(7)
    pdf.cell(0, 7, '分析工具：Claude Competitive Analysis Skill', align='C')
    pdf.ln(14)

    pdf.set_font('zh', '', 8)
    pdf.set_text_color(130, 130, 130)
    pdf.cell(0, 6, '生成日期：2026-06-06', align='C')
    pdf.ln(6)
    pdf.cell(0, 6, '生成者：潘子健 (Zachary Pan)', align='C')

    process_specific(pdf)
    process_concept(pdf)

    pdf.output(OUTPUT)
    print(f'PDF saved to: {OUTPUT}')
    print(f'Pages: {pdf.page_no()}')


if __name__ == '__main__':
    main()
